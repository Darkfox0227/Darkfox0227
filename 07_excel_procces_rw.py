import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import os

os.chdir(r"E:\21_Python_Exersice\07_excel_procces_rw")

wb = openpyxl.load_workbook('example.xlsx')
#type = type(wb)
print(wb.sheetnames) # 查看workbook中的所有worksheets名称:openpyxl.workbook.Workbook.get_sheet_names()

#for sheet in wb:
#    print(sheet.title) # 查看workbook中的所有worksheets名称


print(wb.active) # 開啟活表時出現的那個正在使用的工作表

# get_sheet_by_name 已棄用 改用 wb['']
sheet1 = wb['Sheet1']      #取得工作表Sheet1
#print (wb['Sheet1'])      #查看工作表1
#print (wb['Sheet2'])      #查看工作表2
print (sheet1['A1'])       #取得工作表名稱字串Sheet1
print (sheet1['A1'].value) #從工作表中取得A1儲存格
print (sheet1['B1'].value) #從工作表中取得B1儲存格
print (sheet1['C1'].value) #從工作表中取得C1儲存格

print (sheet1.cell(row=1, column=2))             #取得儲存格名稱字串B1
print (sheet1.cell(row=1, column=2).value)       #取得儲存格B1的Cell物件
#or i in range(1,8,2):
#    print(i, sheet.cell(row=i, column=2).value) #取得儲存格B1的Cell物件

print(sheet1.max_row)    #查工作表最大row數
print(sheet1.max_column) #查工作表最大column數

print(get_column_letter(1))               #欄的數字 -> 字母轉換
print(get_column_letter(2))               #欄的數字 -> 字母轉換
print(get_column_letter(27))              #欄的數字 -> 字母轉換
print(get_column_letter(900))             #欄的數字 -> 字母轉換

print(get_column_letter(sheet1.max_column))
print(column_index_from_string('A'))      #欄的字母 -> 數字轉換
print(column_index_from_string('AA'))     #欄的字母 -> 數字轉換

print(tuple(sheet1['A1':'C1']))            #取得工作表A1到C3區域範圍名稱字串
for rowofcelljects in sheet1['A1':'C1']:          #切片中的每一列
    for cellobj in rowofcelljects:               #讀取每一列中每一個儲存格
        print(cellobj.coordinate, cellobj.value)
    print('----END OF ROW----')



for cellObj in list(sheet1.columns)[1]: #打印单列
    print(cellObj.value)


#for col in sheet1.iter_cols\
#   (min_row = 1, max_row = 7, min_col = 2, max_col = 2):    #印出區塊儲存格範圍
#    for cell in col:
#        print(cell.value)                                   #印出區塊儲存格的值

#sheet = wb.active
#print(sheet.columns[1])
#for cellobj in sheet.columns[1]:
#    print(cellobj.value)
